# pip install openpyxl
# pip install docxtpl
# ternary if operator
# если True условие False
from docxtpl import DocxTemplate
from openpyxl.reader.excel import load_workbook
string = ''
#я решила не вставлят ьэксель файл
wb = load_workbook(filename='')
sheet = wb[''] #wb.active
print(sheet['B4'].value)
rows= sheet.iter_rows(min_row=2, values_only=True)
print(list(rows))
for row in rows:
    num, fio, gender = row
    string = 'Уважаемая' if gender == 'Ж' else string = 'Уважаемый'
    print(string)

string += fio
string += f'\nВаш пригласительный №{num}'
print(string)



temp = 25
print('Тепло') if temp > 25 else print('Прохладно')
if temp > 25:
    print('Тепло')
else:
    print('ПРохладно')

#{{}} jinjia

from docxtpl import DocxTemplate
context = {}
doc = DocxTemplate('template.docx')
for row in rows:
    num, fio, gender = row
    f_name = f'invitation_{num}.docx'
    context = {
        'dear': 'Уважаемый' if gender == 'М' else 'Уважаемая',
        'fio': fio,
        'number': num
    }
    doc.render(context)
    doc.save(f_name)
